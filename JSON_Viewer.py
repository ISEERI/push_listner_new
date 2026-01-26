# JSON_Viewer.py

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import json
import os
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Alignment, Font
from tkcalendar import DateEntry

# Импортируем новые модули
from json_viewer_tables.autoconnect import AutoConnectDisplay
from json_viewer_tables.autoconnect_analysis import AutoConnectAnalysisDisplay
from json_viewer_tables.day_push import DayPushDisplay
import json_viewer_tables.obis_push
from utils import EVENT_DESCRIPTIONS, _format_datetime, classify_day_push_entry, validate_day_push_intervals


class JSONViewer:
    """
    Графический интерфейс для просмотра и анализа DLMS-сообщений,
    сохранённых в JSON-файлах по дням.

    Поддерживает три типа данных:
      - dlms_push: стандартные push-сообщения с OBIS-кодами
      - dlms_day_push: новые push-сообщения с логическим именем и массивом значений
      - autoconnect: служебные сообщения о подключении устройств

    Основные функции:
      - Загрузка данных по дате и типу
      - Отображение в табличной форме
      - Детальный просмотр записи по двойному клику
      - Копирование строк (через контекстное меню или Ctrl+C)
    """

    def __init__(self, parent, initial_load_dir=".", on_load_dir_change=None):
        """
        Инициализация окна просмотра.

        Args:
            parent (tk.Widget): родительский виджет Tkinter.
            initial_load_dir (str): начальная папка для загрузки файлов.
            on_load_dir_change (callable): callback-функция, вызываемая при изменении пути загрузки.
                                          Принимает один аргумент: новый путь (str).
        """
        self.parent = parent
        self.current_display = None
        self.current_data = []  # ← Храним загруженные данные в памяти для детального просмотра
        self.current_file_type = "dlms_push"  # Тип текущего файла (для корректного отображения деталей)
        self.load_data_dir = initial_load_dir  # Папка, откуда загружаются JSON-файлы
        self.export_dir = initial_load_dir  # Папка, куда загружаются файлы Excel
        self.on_load_dir_change = on_load_dir_change
        # Callback для уведомления родителя об изменении пути
        self.setup_ui()  # Создаём пользовательский интерфейс

    def set_load_directory(self, new_dir: str):
        """
        Программно устанавливает новую папку для загрузки и обновляет метку в интерфейсе.
        Используется, например, при загрузке конфигурации извне (DLMSApp).

        Args:
            new_dir (str): Новый путь к папке.
        """
        self.load_data_dir = os.path.normpath(new_dir)
        self.dir_label.config(text=self.load_data_dir)

    def select_load_directory(self):
        """
        Открывает диалог выбора папки для загрузки JSON-файлов.
        Обновляет внутреннее состояние и метку в интерфейсе.
        При успешном выборе вызывает callback on_load_dir_change (если задан).
        """
        selected_dir = filedialog.askdirectory(
            initialdir=self.load_data_dir,
            title="Выберите папку с JSON-файлами для загрузки"
        )
        if selected_dir:
            self.load_data_dir = os.path.normpath(selected_dir)
            self.dir_label.config(text=self.load_data_dir)
            # Уведомляем родительский компонент (DLMSApp) об изменении пути
            if self.on_load_dir_change:
                self.on_load_dir_change(self.load_data_dir)

    def setup_ui(self):

        """Создаёт все элементы управления и таблицу."""
        # === Верхняя панель управления ===
        control_frame = ttk.Frame(self.parent, padding="10")
        control_frame.pack(fill=tk.X)

        # === Выбор папки для загрузки ===
        dir_frame = ttk.Frame(control_frame)
        dir_frame.grid(row=0, column=5, padx=5, sticky=tk.E)  # сместить вправо

        ttk.Label(dir_frame, text="Папка для загрузки:").grid(row=0, column=0, sticky=tk.W, padx=(5, 0))
        self.dir_label = ttk.Label(dir_frame, text=self.load_data_dir, width=30, anchor=tk.W)
        self.dir_label.grid(row=0, column=1, sticky=tk.W, )
        self.dir_button = ttk.Button(dir_frame, text="Выбрать...", command=self.select_load_directory)
        self.dir_button.grid(row=0, column=2, padx=5)

        # === Выбор типа данных (радиокнопки) ===
        ttk.Label(control_frame, text="Тип данных:").grid(row=0, column=0, sticky=tk.W)
        self.file_type_var = tk.StringVar(value="dlms_push")
        file_types = [
            ("Стандартные пуши", "dlms_push"),
            ("Пуши профилей", "dlms_schedule_push"),
            ("Автоконнекты", "autoconnect"),
            ("Анализ автоконнектов", "autoconnect_analysis")
        ]
        for i, (text, value) in enumerate(file_types):
            rb = ttk.Radiobutton(
                control_frame,
                text=text,
                variable=self.file_type_var,
                value=value,
                command=self._on_file_type_change
            )
            rb.grid(row=0, column=i + 1, padx=5, sticky=tk.W)

        # === Поле поиска для анализа автоконнектов ===
        self.search_frame = ttk.Frame(control_frame)
        self.search_frame.grid(row=0, column=4, padx=5, sticky=tk.E)

        self.search_label = ttk.Label(self.search_frame, text="Поиск SN:")
        self.search_label.pack(side=tk.LEFT)
        self.search_var = tk.StringVar()
        self.search_entry = ttk.Entry(self.search_frame, textvariable=self.search_var, width=20)

        # Контекстное меню для строки поиска
        self.search_context_menu = tk.Menu(self.search_entry, tearoff=0)
        self.search_context_menu.add_command(label="Вставить", command=self._search_paste)
        self.search_context_menu.add_command(label="Копировать", command=self._search_copy)
        self.search_context_menu.add_command(label="Вырезать", command=self._search_cut)
        self.search_context_menu.add_separator()
        self.search_context_menu.add_command(label="Выделить всё", command=self._search_select_all)

        # Привязка ПКМ
        self.search_entry.bind("<Button-3>", self._show_search_context_menu)
        self.search_entry.pack(side=tk.LEFT, padx=(5, 0))
        self.search_var.trace_add("write", self._on_search_change)

        # Изначально скрываем поле поиска
        self.search_frame.grid_remove()

        # === Выбор диапазона дат ===
        ttk.Label(control_frame, text="Период").grid(row=1, column=0, sticky=tk.W, pady=(10, 0))

        # Создаем Frame для "С:" + DateEntry
        start_frame = ttk.Frame(control_frame)
        start_frame.grid(row=1, column=1, sticky=tk.W, padx=(5, 0), pady=(10, 0))
        ttk.Label(start_frame, text="с:").pack(side=tk.LEFT)
        self.start_date = DateEntry(
            start_frame,
            width=12,
            background='black',
            foreground='white',
            borderwidth=2,
            date_pattern='dd.mm.yyyy',
            locale='ru_RU'
        )
        self.start_date.set_date(datetime.now())
        self.start_date.pack(side=tk.LEFT, padx=(5, 0))

        # Создаем Frame для "По:" + DateEntry
        end_frame = ttk.Frame(control_frame)
        end_frame.grid(row=1, column=2, sticky=tk.W, padx=(10, 0), pady=(10, 0))
        ttk.Label(end_frame, text="по:").pack(side=tk.LEFT)
        self.end_date = DateEntry(
            end_frame,
            width=12,
            background='black',
            foreground='white',
            borderwidth=2,
            date_pattern='dd.mm.yyyy',
            locale='ru_RU'
        )
        self.end_date.set_date(datetime.now())
        self.end_date.pack(side=tk.LEFT, padx=(5, 0))

        # Кнопки управления
        button_frame = ttk.Frame(control_frame)
        button_frame.grid(row=1, column=4, pady=(10, 0), padx=5)

        self.load_btn = ttk.Button(button_frame, text="Загрузить", command=self.load_json_data)
        self.load_btn.grid(row=0, column=0, padx=(0, 5))  # небольшой отступ между кнопками

        self.refresh_btn = ttk.Button(button_frame, text="Обновить", command=self.load_json_data)
        self.refresh_btn.grid(row=0, column=1, padx=(5, 0))
        self.between_label = ttk.Label(button_frame, text="", width=30, anchor=tk.W)
        self.between_label.grid(row=0, column=2, sticky=tk.W, )

        # === Фрейм для экспорта в Excel ===
        export_frame = ttk.Frame(control_frame)
        export_frame.grid(row=1, column=5, pady=(10, 0), padx=5, sticky=tk.E)  # сместить вправо

        # Кнопка экспорта
        self.export_btn = ttk.Button(export_frame, text="Экспорт в Excel", command=self.export_table_to_excel)
        self.export_btn.grid(row=0, column=0, padx=(0, 10))

        # Метка "Папка Excel:"
        ttk.Label(export_frame, text="Папка Excel:").grid(row=0, column=1, sticky=tk.W, padx=(0, 5))

        # Текущий путь
        self.export_dir_label = ttk.Label(export_frame, text=self.export_dir, width=30, anchor=tk.W)
        self.export_dir_label.grid(row=0, column=2, sticky=tk.W, padx=(0, 5))

        # Кнопка выбора
        ttk.Button(export_frame, text="Выбрать...", command=self.select_export_directory).grid(row=0, column=3,
                                                                                               padx=(0, 5))
        # === Таблица для отображения данных ===
        table_frame = ttk.Frame(self.parent)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        self.tree = ttk.Treeview(table_frame, show="headings")
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        hsb.pack(side=tk.BOTTOM, fill=tk.X)

        # === Контекстное меню и горячие клавиши ===
        self.tree_context_menu = tk.Menu(self.tree, tearoff=0)
        self.tree_context_menu.add_command(label="Копировать строку", command=self.copy_tree_row)
        self.tree.bind("<Button-3>", self.show_tree_context_menu)  # ПКМ

        # Стандартное событие копирования (Cmd+C на macOS, Ctrl+C на Windows/Linux)
        self.tree.bind("<<Copy>>", lambda e: self.copy_tree_row())

        # Обработка нажатий клавиш на всём окне (для Ctrl+C через keycode=67)
        self.parent.bind_all("<Key>", self._on_key_press)

        # Двойной клик — открыть детали
        self.tree.bind("<Double-1>", self._on_double_click)

    def _show_search_context_menu(self, event):
        """Показывает контекстное меню для строки поиска."""
        self.search_context_menu.tk_popup(event.x_root, event.y_root)

    def _search_paste(self):
        try:
            text = self.parent.clipboard_get()
            self.search_entry.insert(tk.INSERT, text)
        except tk.TclError:
            pass  # Буфер пуст

    def _search_copy(self):
        try:
            selected_text = self.search_entry.selection_get()
            self.parent.clipboard_clear()
            self.parent.clipboard_append(selected_text)
        except tk.TclError:
            pass  # Ничего не выделено

    def _search_cut(self):
        try:
            selected_text = self.search_entry.selection_get()
            self.parent.clipboard_clear()
            self.parent.clipboard_append(selected_text)
            self.search_entry.delete(tk.SEL_FIRST, tk.SEL_LAST)
        except tk.TclError:
            pass  # Ничего не выделено

    def _search_select_all(self, event=None):
        self.search_entry.select_range(0, tk.END)
        self.search_entry.icursor(tk.END)  # Переместить курсор в конец
        return "break"  # Предотвратить дальнейшую обработку

    def _on_key_press(self, event):
        """
        Обрабатывает глобальное нажатие клавиш.
        Реализует копирование строки при Ctrl+C (keycode=67), если фокус на Treeview.
        """
        # Проверяем: нажат ли Ctrl (state & 0x4) и клавиша C (keycode=67)
        if event.state & 0x4 and event.keycode == 67:
            if str(self.parent.focus_get()) == str(self.tree):
                self.copy_tree_row()
                return "break"  # Блокируем дальнейшую обработку
        elif event.state & 0x4 and event.keycode == 55:
            if str(self.parent.focus_get()) == str(self.tree):
                self._search_paste()
                return "break"
    def _on_double_click(self, event):
        """
        Обрабатывает двойной клик по строке таблицы.
        Открывает модальное окно с детальной информацией о записи.
        """
        item = self.tree.selection()
        if not item:
            return
        item = item[0]
        try:
            index = self.tree.index(item)
            if 0 <= index < len(self.current_data):
                entry = self.current_data[index]
                self._show_detail_window(entry, self.current_file_type)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось открыть детали:\n{e}")

    def _on_file_type_change(self):
        """Показывает/скрывает поле поиска в зависимости от выбранного типа."""
        if self.file_type_var.get() == "autoconnect_analysis":
            self.search_frame.grid()  # Показать
        else:
            self.search_frame.grid_remove()  # Скрыть

    def export_table_to_excel(self):
        """Экспортирует данные в Excel с учётом типа отображения."""
        file_type = self.file_type_var.get()

        if file_type == "autoconnect_analysis":
            self._export_autoconnect_analysis_to_excel()
        else:
            self._export_standard_table_to_excel()

    def _export_standard_table_to_excel(self):
        """Стандартный экспорт для плоских таблиц."""
        all_items = self.tree.get_children()
        if not all_items:
            messagebox.showinfo("Информация", "Таблица пуста")
            return

        # Формируем имя файла
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"export_{timestamp}.xlsx"
        filepath = os.path.join(self.export_dir, filename)

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Данные"

            headers = [self.tree.heading(col, option="text") for col in self.tree["columns"]]
            ws.append(headers)

            for item_id in all_items:
                values = self.tree.item(item_id, "values")
                ws.append(list(values))

            self._apply_excel_styles(wb)
            wb.save(filepath)
            messagebox.showinfo("Успешно", f"Файл сохранён:\n{filepath}")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось экспортировать в Excel:\n{e}")

    def _export_autoconnect_analysis_to_excel(self):
        """Специальный экспорт для анализа автоконнектов с группировкой по SN."""
        all_items = self.tree.get_children()
        if not all_items:
            messagebox.showinfo("Информация", "Таблица пуста")
            return

        # Формируем имя файла
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"export_autoconnect_{timestamp}.xlsx"
        filepath = os.path.join(self.export_dir, filename)

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Анализ автоконнектов"

            # Заголовки (включая колонку SN)
            headers = ["Счётчик (SN)", "Получено", "IP", "Порт", "Валидация"]
            ws.append(headers)

            for parent_id in all_items:
                sn = self.tree.item(parent_id, "text")
                # Получаем все дочерние элементы
                child_items = self.tree.get_children(parent_id)
                if not child_items:
                    # Если нет дочерних элементов, добавляем пустую строку
                    ws.append([sn, "", "", "", ""])
                else:
                    for child_id in child_items:
                        values = self.tree.item(child_id, "values")
                        # values содержит: (received_at, ip, port, validation)
                        row = [sn] + list(values)
                        ws.append(row)

            self._apply_excel_styles(wb)
            wb.save(filepath)
            messagebox.showinfo("Успешно", f"Файл сохранён:\n{filepath}")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось экспортировать в Excel:\n{e}")

    def _apply_excel_styles(self, wb):
        """Применяет стили к Excel-файлу."""
        ws = wb.active
        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center')

        for cell in ws[1]:
            cell.font = bold_font
            cell.alignment = center_alignment

        for col_num, column_cells in enumerate(ws.columns, 1):
            max_length = 0
            for cell in column_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)  # Ограничиваем ширину
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = adjusted_width

    def select_export_directory(self):
        """Открывает диалог выбора папки для экспорта Excel-файлов."""
        selected_dir = filedialog.askdirectory(
            initialdir=self.export_dir,
            title="Выберите папку для сохранения Excel-файлов"
        )
        if selected_dir:
            self.export_dir = os.path.normpath(selected_dir)
            self.export_dir_label.config(text=self.export_dir)

    def _on_search_change(self, *args):
        """Обновляет отображение при изменении поиска (только для анализа автоконнектов)."""
        if self.file_type_var.get() == "autoconnect_analysis" and self.current_display:
            # Передаём текущие данные и фильтр
            search_filter = self.search_var.get()
            self.current_display.display(self.current_data, search_filter=search_filter)

    def _show_detail_window(self, entry, file_type):
        """
        Открывает модальное окно с полной информацией о выбранной записи.
        Формат зависит от типа файла.

        Args:
            entry (dict): запись из JSON-файла
            file_type (str): тип данных ('dlms_push', 'dlms_day_push', 'autoconnect')
        """
        title = "Детали записи"
        detail_win = tk.Toplevel(self.parent)
        detail_win.title(title)
        detail_win.geometry("600x400")

        # Текстовое поле для отображения данных
        text = tk.Text(detail_win, wrap=tk.WORD, font=("Consolas", 10))
        text.pack(fill=tk.BOTH, expand=True)

        # Контекстное меню для текстового поля
        context_menu = tk.Menu(text, tearoff=0)
        context_menu.add_command(label="Копировать", command=lambda: self._copy_text_selection(text))
        text.bind("<Button-3>", lambda e: context_menu.tk_popup(e.x_root, e.y_root))

        # Обработка Ctrl+C внутри текстового поля
        def on_text_key(event):
            if event.state & 0x4 and event.keycode == 67:
                self._copy_text_selection(text)
                return "break"

        text.bind("<Key>", on_text_key)

        # === Формирование содержимого в зависимости от типа ===
        if file_type in ("autoconnect", "autoconnect_analysis"):
            # Отображение данных автоконнекта
            received_at = _format_datetime(entry.get("received_at", ""))
            sn = entry.get("sn", "")
            ip = entry.get("ip", "")
            port = entry.get("port", "")
            text.insert("1.0", f"Получено: {received_at}\nSN: {sn}\nIP: {ip}\nПорт: {port}")
        elif file_type == "dlms_schedule_push":
            # Отображение day-push с валидацией профиля
            received_at = _format_datetime(entry.get("received_at", ""))
            invoke_id = entry.get("invoke_id", "")
            logical_name = entry.get("logical_name", "")

            profile_type = classify_day_push_entry(entry)
            profile_label = {"daily": "Суточный профиль", "half_hourly": "Профиль энергии 1"}.get(profile_type,
                                                                                                  "Неизвестен")
            is_valid = validate_day_push_intervals(entry, profile_type)
            valid_status = "Корректно" if is_valid else "ОШИБКА: нарушена периодичность!"

            text.insert("1.0",
                        f"Получено: {received_at}\nInvoke ID: {invoke_id}\nЛогическое имя: {logical_name}\nТип профиля: {profile_label}\nВалидность: {valid_status}\n\n")

            for record in entry.get("data", []):
                ts = _format_datetime(record.get("timestamp", ""))
                vals = ", ".join(str(v) for v in record.get("values", []))
                text.insert(tk.END, f"{ts}: [{vals}]\n")
        else:  # dlms_push
            # Отображение стандартных OBIS-push сообщений
            received_at = _format_datetime(entry.get("received_at", ""))
            invoke_id = entry.get("invoke_id", "")
            text.insert("1.0", f"Получено: {received_at}\nInvoke ID: {invoke_id}\n\n")

            for record in entry.get("records", []):
                obis = record.get("obis", "")
                value_obj = record.get("value", {})

                # === СПЕЦИАЛЬНАЯ ОБРАБОТКА ДЛЯ 0.0.97.98.0.255 ===
                if obis == "0.0.97.98.0.255":
                    value_str = json_viewer_tables.obis_push._format_value(value_obj)
                    # Получаем hex-значение из парсера
                    hex_val = value_obj.get("value", "")
                    if len(hex_val) == 8:  # 4 байта = 8 hex-символов
                        try:
                            # Преобразуем в целое число (big-endian)
                            bitmask = int(hex_val, 16)
                            # Расшифровываем события
                            events = []
                            for bit in range(32):
                                if bitmask & (1 << bit):
                                    desc = EVENT_DESCRIPTIONS.get(bit, f"Неизвестное событие {bit}")
                                    events.append(f"Бит {bit:2d}: {desc}")

                            if events:
                                line = f"{obis}: {value_str}\n Активные события:\n"
                                for event in events:
                                    line += f"  • {event}\n"
                            else:
                                line = f"{obis}: Нет активных событий\n"
                        except ValueError:
                            line = f"{obis}: Ошибка парсинга значения '{hex_val}'\n"
                    else:
                        line = f"{obis}: Некорректная длина значения ({len(hex_val)} символов)\n"
                else:
                    # Обычная обработка
                    value_str = json_viewer_tables.obis_push._format_value(value_obj)
                    line = f"{obis}: {value_str}\n"

                text.insert(tk.END, line)

        text.configure(state="disabled")  # Только для чтения

    def _copy_text_selection(self, text_widget):
        """Копирует выделенный текст из Text-виджета в буфер обмена."""
        try:
            selected_text = text_widget.get(tk.SEL_FIRST, tk.SEL_LAST)
            text_widget.clipboard_clear()
            text_widget.clipboard_append(selected_text)
        except tk.TclError:
            pass  # Ничего не выделено — игнорируем

    def show_tree_context_menu(self, event):
        """Отображает контекстное меню при правом клике по строке таблицы."""
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)
            try:
                self.tree_context_menu.tk_popup(event.x_root, event.y_root)
            finally:
                self.tree_context_menu.grab_release()

    def copy_tree_row(self):
        """Копирует всю строку таблицы (все колонки, разделённые табуляцией) в буфер обмена."""
        selected = self.tree.selection()
        if not selected:
            return
        values = self.tree.item(selected[0], "values")
        row_text = "\t".join(str(v) for v in values)
        self.parent.clipboard_clear()
        self.parent.clipboard_append(row_text)

    def load_json_data(self):
        file_type = self.file_type_var.get()
        try:
            start_dt = self.start_date.get_date()
            end_dt = self.end_date.get_date()
            if start_dt > end_dt:
                messagebox.showerror("Ошибка", "Начальная дата не может быть позже конечной!")
                return
        except Exception as e:
            messagebox.showerror("Ошибка", f"Неверный формат даты:\n{e}")
            return
        # Определяем, из какого файла читать данные
        if file_type == "autoconnect_analysis":
            data_file_prefix = "autoconnect"  # ← Источник данных тот же!
        else:
            data_file_prefix = file_type

        all_data = []
        current_date = start_dt
        while current_date <= end_dt:
            filename = f"{data_file_prefix}_{current_date.strftime('%Y-%m-%d')}.json"
            full_path = os.path.join(self.load_data_dir, filename)
            if os.path.exists(full_path):
                try:
                    with open(full_path, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                        if isinstance(data, list):
                            all_data.extend(data)
                except Exception as e:
                    messagebox.showwarning("Предупреждение", f"Не удалось прочитать {filename}:\n{e}")
            current_date += timedelta(days=1)

        if not all_data:
            messagebox.showinfo("Информация", "Нет данных за выбранный период")
            return

        self.current_data = all_data
        self.current_file_type = file_type

        # === ДЕЛЕГИРОВАНИЕ ОТОБРАЖЕНИЯ ===
        if file_type == "autoconnect":
            display = AutoConnectDisplay(self.tree, self.parent)
        elif file_type == "autoconnect_analysis":
            display = AutoConnectAnalysisDisplay(self.tree, self.parent)
        elif file_type == "dlms_schedule_push":
            display = DayPushDisplay(self.tree, self.parent)
        else:  # dlms_push
            display = json_viewer_tables.obis_push.ObisPushDisplay(self.tree, self.parent)

        if file_type == "autoconnect_analysis":
            # Передаём текущий фильтр
            search_filter = self.search_var.get()
            display.display(all_data, search_filter=search_filter)
        else:
            display.display(all_data)
